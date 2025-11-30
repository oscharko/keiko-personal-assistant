"""
Export functionality for the Ideas Hub module.

Provides export capabilities in multiple formats:
- Excel (XLSX) with all idea data and scores
- PDF summary report
- CSV for data analysis
"""

import csv
import io
import logging
from datetime import datetime
from typing import Any

from .models import Idea

logger = logging.getLogger(__name__)


class IdeasExporter:
    """
    Handles exporting ideas to various formats.

    Supports Excel, PDF, and CSV exports with configurable fields.
    """

    # Default fields to include in exports
    DEFAULT_FIELDS = [
        "ideaId",
        "title",
        "description",
        "summary",
        "status",
        "department",
        "submitterId",
        "createdAt",
        "updatedAt",
        "impactScore",
        "feasibilityScore",
        "recommendationClass",
        "clusterLabel",
        "tags",
    ]

    def __init__(self):
        """Initialize the exporter."""
        pass

    def _format_timestamp(self, timestamp: int | None) -> str:
        """Convert millisecond timestamp to ISO format string."""
        if not timestamp:
            return ""
        try:
            dt = datetime.fromtimestamp(timestamp / 1000)
            return dt.strftime("%Y-%m-%d %H:%M:%S")
        except (ValueError, OSError):
            return ""

    def _format_list(self, items: list | None) -> str:
        """Convert list to comma-separated string."""
        if not items:
            return ""
        return ", ".join(str(item) for item in items)

    def _idea_to_row(self, idea: Idea, fields: list[str]) -> dict[str, Any]:
        """
        Convert an idea to a row dictionary for export.

        Args:
            idea: The idea to convert.
            fields: List of field names to include.

        Returns:
            Dictionary with field values.
        """
        idea_dict = idea.to_dict()
        row = {}

        for field in fields:
            value = idea_dict.get(field)

            # Format special fields
            if field in ("createdAt", "updatedAt"):
                value = self._format_timestamp(value)
            elif field == "tags":
                value = self._format_list(value)
            elif field == "affectedProcesses":
                value = self._format_list(value)
            elif field == "targetUsers":
                value = self._format_list(value)
            elif isinstance(value, list):
                value = self._format_list(value)
            elif value is None:
                value = ""

            row[field] = value

        return row

    def export_to_csv(
        self,
        ideas: list[Idea],
        fields: list[str] | None = None,
    ) -> str:
        """
        Export ideas to CSV format.

        Args:
            ideas: List of ideas to export.
            fields: List of field names to include (default: DEFAULT_FIELDS).

        Returns:
            CSV content as string.
        """
        if not fields:
            fields = self.DEFAULT_FIELDS

        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=fields)
        writer.writeheader()

        for idea in ideas:
            row = self._idea_to_row(idea, fields)
            writer.writerow(row)

        return output.getvalue()

    def export_to_excel(
        self,
        ideas: list[Idea],
        fields: list[str] | None = None,
    ) -> bytes:
        """
        Export ideas to Excel format.

        Args:
            ideas: List of ideas to export.
            fields: List of field names to include (default: DEFAULT_FIELDS).

        Returns:
            Excel file content as bytes.
        """
        if not fields:
            fields = self.DEFAULT_FIELDS

        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.warning("openpyxl not installed, falling back to CSV")
            csv_content = self.export_to_csv(ideas, fields)
            return csv_content.encode("utf-8")

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ideas"

        # Header style
        header_font = Font(bold=True, color="000000")
        header_fill = PatternFill(
            start_color="DCFF4A",
            end_color="DCFF4A",
            fill_type="solid",
        )
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Write headers
        for col_idx, field in enumerate(fields, 1):
            cell = ws.cell(row=1, column=col_idx, value=field)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Write data rows
        for row_idx, idea in enumerate(ideas, 2):
            row_data = self._idea_to_row(idea, fields)
            for col_idx, field in enumerate(fields, 1):
                ws.cell(row=row_idx, column=col_idx, value=row_data.get(field, ""))

        # Auto-adjust column widths
        for col_idx, field in enumerate(fields, 1):
            column_letter = get_column_letter(col_idx)
            max_length = len(field)
            for row in ws.iter_rows(
                min_row=2,
                max_row=len(ideas) + 1,
                min_col=col_idx,
                max_col=col_idx,
            ):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def export_summary_report(
        self,
        ideas: list[Idea],
        title: str = "Ideas Summary Report",
    ) -> str:
        """
        Generate a text summary report of ideas.

        Args:
            ideas: List of ideas to summarize.
            title: Report title.

        Returns:
            Report content as string.
        """
        lines = []
        lines.append("=" * 60)
        lines.append(title)
        lines.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append("=" * 60)
        lines.append("")

        # Summary statistics
        total = len(ideas)
        if total == 0:
            lines.append("No ideas to report.")
            return "\n".join(lines)

        # Count by status
        status_counts: dict[str, int] = {}
        for idea in ideas:
            status = idea.status.value if idea.status else "unknown"
            status_counts[status] = status_counts.get(status, 0) + 1

        # Count by recommendation class
        rec_counts: dict[str, int] = {}
        for idea in ideas:
            rec = idea.recommendation_class or "unclassified"
            rec_counts[rec] = rec_counts.get(rec, 0) + 1

        # Average scores
        impact_scores = [i.impact_score for i in ideas if i.impact_score > 0]
        feasibility_scores = [i.feasibility_score for i in ideas if i.feasibility_score > 0]

        lines.append("SUMMARY STATISTICS")
        lines.append("-" * 40)
        lines.append(f"Total Ideas: {total}")
        lines.append("")

        lines.append("By Status:")
        for status, count in sorted(status_counts.items()):
            lines.append(f"  - {status}: {count}")
        lines.append("")

        lines.append("By Recommendation:")
        for rec, count in sorted(rec_counts.items()):
            lines.append(f"  - {rec}: {count}")
        lines.append("")

        if impact_scores:
            avg_impact = sum(impact_scores) / len(impact_scores)
            lines.append(f"Average Impact Score: {avg_impact:.1f}")
        if feasibility_scores:
            avg_feasibility = sum(feasibility_scores) / len(feasibility_scores)
            lines.append(f"Average Feasibility Score: {avg_feasibility:.1f}")
        lines.append("")

        # Top ideas by impact
        lines.append("TOP 5 IDEAS BY IMPACT")
        lines.append("-" * 40)
        top_impact = sorted(ideas, key=lambda x: x.impact_score, reverse=True)[:5]
        for i, idea in enumerate(top_impact, 1):
            lines.append(f"{i}. {idea.title}")
            lines.append(f"   Impact: {idea.impact_score}, Feasibility: {idea.feasibility_score}")
            lines.append(f"   Recommendation: {idea.recommendation_class or 'N/A'}")
            lines.append("")

        # Quick wins
        quick_wins = [i for i in ideas if i.recommendation_class == "quick_win"]
        if quick_wins:
            lines.append("QUICK WINS")
            lines.append("-" * 40)
            for idea in quick_wins[:5]:
                lines.append(f"- {idea.title}")
                lines.append(f"  Impact: {idea.impact_score}, Feasibility: {idea.feasibility_score}")
            lines.append("")

        lines.append("=" * 60)
        lines.append("End of Report")
        lines.append("=" * 60)

        return "\n".join(lines)

